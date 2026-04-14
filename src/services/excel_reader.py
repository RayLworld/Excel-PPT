from openpyxl import load_workbook


def read_excel_records(excel_path, header_row):
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        sheet = workbook.active

        header_cells = list(
            sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )
        if not header_cells:
            raise ValueError("未读取到表头行，请检查Excel表头行号")

        raw_headers = list(header_cells[0])
        data_rows = [
            list(row)
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True)
        ]
    finally:
        workbook.close()

    valid_rows = [row for row in data_rows if any(cell is not None for cell in row)]
    max_col_count = (
        max([len(raw_headers)] + [len(row) for row in valid_rows])
        if valid_rows
        else len(raw_headers)
    )
    normalized_rows = [
        row + [None] * (max_col_count - len(row))
        for row in valid_rows
    ]
    normalized_headers = raw_headers + [None] * (max_col_count - len(raw_headers))

    keep_col_indices = [
        col_idx
        for col_idx in range(max_col_count)
        if any(row[col_idx] is not None for row in normalized_rows)
    ]

    column_names = []
    for idx, col_idx in enumerate(keep_col_indices):
        header_value = normalized_headers[col_idx]
        if header_value is None:
            column_names.append(f"列_{idx + 1}")
        else:
            column_names.append(str(header_value))

    records = []
    for row in normalized_rows:
        record = {
            column_names[idx]: row[col_idx]
            for idx, col_idx in enumerate(keep_col_indices)
        }
        records.append(record)

    return records, column_names
